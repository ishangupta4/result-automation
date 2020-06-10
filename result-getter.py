from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from time import sleep
from openpyxl import Workbook, load_workbook

wb = load_workbook("") #Name of the excel workbook
ws = wb[""]	#worksheet to read
wr = wb[""]	#worksheet to write
enroll = []
faculty = []
for i in range(64):
    enroll.append(ws["C" + str(i + 4)].value)
    faculty.append(ws["B" + str(i + 4)].value)

driver = webdriver.Chrome()
driver.get("")				#url of the site
WebDriverWait(driver, 15).until(
    EC.presence_of_element_located((By.LINK_TEXT, ""))		#link text you want to identify 
)

for i in range(64):
    driver.get("")		#url of the site
    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.LINK_TEXT, ""))	#link text you want to identify
    )
    driver.find_element_by_xpath('//*[@id="myNavbar"]/ul/li[3]/a').click()
    driver.find_element_by_xpath('//*[@id="myNavbar"]/ul/li[3]/ul/li[14]/a').click()
    driver.find_element_by_id("facno_input").send_keys(faculty[i])
    driver.find_element_by_name("en").send_keys(enroll[i])
    driver.find_element_by_id("att_submit").click()
    WebDriverWait(driver, 5).until(
        EC.presence_of_element_located(
            (By.LINK_TEXT, "Click here to check your equivalent percentage...")
        )
    )
    name = driver.find_element_by_xpath(
        "/html/body/div[2]/div/div/table[2]/tbody/tr[2]/td[3]"
    ).text
    spi = driver.find_element_by_xpath(
        "/html/body/div[2]/div/div/table[2]/tbody/tr[2]/td[5]"
    ).text
    print(str(name) + " : " + str(spi))
    wr["A" + str(i + 1)].value = name
    wr["B" + str(i + 1)].value = spi
    wb.save("")		#workbook to save

driver.quit()
